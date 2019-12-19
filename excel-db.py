import openpyxl

import sqlite3

def readData(xl_file_name):

    """
        Function to read all the data that is present in the excel workbook.
        This returns a dictionary consisting of all the following:
            Number of rows
            Table name
            Column names
            Row data
    """

    wb = openpyxl.load_workbook(xl_file_name + ".xlsx")

    # To get all the worksheets
    sheets = wb.sheetnames
    # To select the individual sheet
    sheet = wb[sheets[0]]

    # Number of rows and columns
    no_of_row = sheet.max_row + 1
    no_of_col = sheet.max_column + 1

    # The start values for the rows and columns respectively
    row_start = 3
    col_start = 2

    # Loop variables
    row_counter = 1
    col_counter = 1

    # List for storing the names of the columns
    col_names = list()

    # Variable to hold the cell where the table name is stored
    first_cell = "A1"

    # Getting the table name
    table_name = sheet[first_cell].value

    # Getting the column names
    for col_counter in range(1, no_of_col):
        col_names.append(sheet.cell(2, col_counter).value)

    # List for storing the data of each rows
    row_data = list()

    # Loop to iterate over each row
    for row_counter in range(row_start, no_of_row):
        # Temporary list to store the data of the individual row and then append it to the row_data
        # so that temp_row_data can be rewritten
        temp_row_data = list()
        # Loop to iterate over the columns
        for col_counter in range(1, no_of_col):
            # Appending the data of the particular row into the temporary list
            temp_row_data.append(sheet.cell(row_counter, col_counter).value)
        # Appending the temp_row_data to row_data
        row_data.append(temp_row_data)

    # Subtracting 1 from the number of columns and rows because they were incremented by one for the loop
    
    result = dict()

    result = {
        'rows': no_of_row - 1,
        'table_name': table_name,
        'col_names': col_names,
        'row_data': row_data
    }

    return result

def dtypeDB(col_name):
   
    """
        Funtion used to return the datatype of the data in the worksheet to its
        equivalent datatype in sqlite3
    """

    if type(col_name) == int:
        return 'integer'
    elif type(col_name) == float:
        return 'real'
    elif type(col_name) == str:
        return 'text'
    elif type(col_name) == None:
        return 'none'
    else:
        return 'blob'

def createStatement(res):

    """
        This function returns statement for creating the table in DB
    """

    # Var which holds the index in the row data which determines the type of data which
    # all elements of the respective column will be cast into 
    index = 0

    # This var holds the create table statement in which the table is name is formatted initially
    create_table_statement = "create table {}(".format(res['table_name'])

    # Loop to append the column name along with the datatype of the data in the worksheet
    for col_name in res["col_names"]:
        create_table_statement += str(col_name) + " " + dtypeDB(res['row_data'][0][index])
        index += 1

        # Not adding comma when all the column names are appended
        if col_name != res['col_names'][-1]:
            create_table_statement += ", "

    # Appending the closing bracket to the create statement string
    create_table_statement += (')')

    return create_table_statement

def insertStatements(res):

    """
        This function returns the set of statements for inserting the data into the DB
    """

    # List for holding all the values that are to be inserted
    insert_data_commands = list()

    # Loop to iterate over all the row data lists
    for cell_list in res['row_data']:
        # Formatting in the table name to var insert_data_statement
        # This is a temp var that holds each of the commands and is overrwritten during each iteration
        insert_data_statement = "insert into {} values(".format(res['table_name'])
        # Looping over all the data that are to be inserted into the database
        for cell_data in cell_list:

            if type(cell_data) == str:
                temp = cell_data
                cell_data = "'" + cell_data + "'"
                insert_data_statement += str(cell_data)
                cell_data = temp

            else:
                insert_data_statement += str(cell_data)
            
            # Not adding comma when all the column names are appended
            if cell_data != cell_list[-1]:
                insert_data_statement += ', '

        # Appending the closing bracket to the create statement string
        insert_data_statement += (')')

        # Appending each insert query to the list insert_data_commands
        insert_data_commands.append(insert_data_statement)
    
    return insert_data_commands

def dataBaseInteract(db_name, res, create_statement, insert_statements):
    
    """
        This function interacts with the DB and performs the following 
            Connecting to DB or creating a new DB
            Creating a table
            Inserting the data into the table
    """

    try:
        conn = sqlite3.connect(db_name + ".db")

        c = conn.cursor()

        # Creating the table in the DB
        c.execute(create_statement)
        conn.commit()

        # Inserting the data into the table
        for query in insert_statements:
            c.execute(query)
            conn.commit()

        conn.commit()

        c.execute("select * from {}".format(res['table_name']))

        print("\nData inserted into the DB are as follows:")
        print(c.fetchall())
        print()

    except Exception:
        print("Something went wrong...")

    else:
        print("Converted into DB successfully")

    finally:
        conn.close()

def run():

    """
        This is the main function which runs all the functions and performs the following tasks
            Read the excel file name to be read from the user
            Read the DB file name from the user
            Call function which reads the excel file
            Call function which creates the table creation query
            Call function which creates the table insertion queries
            Call function which creates DB file and inserts the data into the DB
    """

    xl_file_name = input("Enter the name of the excel file which is to be converted (do not type the extension): ")

    db_name = input("Enter the name of the DB file: ")

    print("Reading the excel file")
    res = readData(xl_file_name)

    print("Creating the create query statement")
    create_statement = createStatement(res)

    print("Creating the insert query statements")
    insert_statements = insertStatements(res)

    print("Inserting into the DB")
    dataBaseInteract(db_name, res, create_statement, insert_statements)

# Calling the main function
run()